from openpyxl import load_workbook

stepeni_wb = load_workbook(filename="stepeni.xlsx")
find_wb = load_workbook(filename="find.xlsx")
for sheet in find_wb:
    for row in sheet.iter_rows(min_row=1, max_col=1, values_only=True):
        for cell in row:
            surname = cell[:cell.find(" ")]
            for sheet_stepeni in stepeni_wb:
                for row_stepeni in sheet_stepeni.iter_rows(min_row=1, min_col=2, max_col=2, values_only=True):
                    if row_stepeni[0]:
                        if surname in row_stepeni[0]:
                            print(row_stepeni[0])